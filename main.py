from openpyxl import Workbook, load_workbook
import pandas
import time
# Data functions

def get_sheets_names():
    """returing  array of sheet names in data.xlsx."
    """
    try:
        file = pandas.ExcelFile('data.xlsx')
    except:
        print("Error reading xlsx file")

    sheet_names = file.sheet_names
    return sheet_names

def get_points_from_sheet(sheet_name):
    """returing points from [sheet_name] in data.xlsx. returning array of sorted points,
    sorted by x value. if x value is equal, sort by y value"""
    try:
        DataFrame = pandas.read_excel("data.xlsx",sheet_name=sheet_name)
    except:
        print("Error reading xlsx file")

    num_of_points = DataFrame.shape[0]
    points = [(DataFrame['X'][i], DataFrame['Y'][i]) for i in range(num_of_points)]
    sorted_points = sorted(points, key=lambda tup: (tup[0], tup[1]))
    return sorted_points,points

def write_output(points,weighted_points_map, squared_weighted_points_map,sheet_name):
    """creating output.xlsx file with n points."""

    try:
        wb = load_workbook('output.xlsx')
    except:
        wb = Workbook()

    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])
    
    ws = wb.create_sheet(title=sheet_name)

    ws['A1'] = 'X'
    ws['B1'] = 'Y'
    ws['C1'] = 'W'
    ws['D1'] = 'SQ'
    for i,point in enumerate(points):
        x,y = point[0],point[1]
        w = weighted_points_map.get(point, 1)
        sq = squared_weighted_points_map.get(point, 1)
        ws[f"A{i+2}"] = x
        ws[f"B{i+2}"] = y
        ws[f"C{i+2}"] = w
        ws[f"D{i+2}"] = sq

    
    # try:
    wb.save("output.xlsx")
    # except:
    #     print("Error occured in wb.save()")

# Algorithm functions

# Calculate W
def get_W(points, weights_map,squared):
    n = len(points)
    dp = [0] * n

    weighted_points = []
    for point in points:
        w = weights_map.get(point, 1)
        weighted_points.append((point[0], point[1], w))
    for i in range(n - 1, -1, -1):
        xi, yi, wi = weighted_points[i]
        dp[i] = wi

        for j in range(i + 1, n):
            xj, yj, _ = weighted_points[j]

            if squared == False:     
                if xi < xj and yi < yj:
                    if dp[i] < dp[j] + wi:
                        dp[i] = dp[j] + wi
            else:
                if xi < xj and yi < yj:
                    if dp[i] < dp[j] + (wi) ** 2:
                        dp[i] = dp[j] + (wi) ** 2

    return max(dp), weighted_points

# Regular sum of weights functions
def get_set_of_points_on_heaviest_chains(points, weights_map):
    """returns a map: KEY = point (x,y) , VALUE = ( _w , set) : _w is the weight of heaviest chain starting at point. set is a set of all points on _w weight chains starting from point."""

    print("1. get_set_of_points_on_heaviest_chains()")

    # KEY = point , VALUE = (_w, set()) . this is the returned map.
    point_set_map = {}


    # Iterate from end to start.
    # for each point (point_i) calculate the following:
    # 1. iset = set of all points that are on one of the heaviest weight chains starting from point_i.
    # 2. _w = weight of the heaviest chains.

    for i in range(len(points) - 1, -1, -1):
        # for printing.
        perc = 100 - (i / len(points)) * 100

        point_i = (points[i][0], points[i][1])
        iset = set()
        _w = 0

        # find the max _w belongs to any comparable bigger point (point_j)
        for j in range(i + 1, len(points)):
            point_j = (points[j][0], points[j][1])
            if point_j[0] > point_i[0] and point_j[1] > point_i[1]:
                # point_j is a comparable bigger point!
                if (value := point_set_map.get(points[j])) is not None:
                    # point_j has been found.
                    jweight = value[0]
                    _w = max(_w, jweight)

        # now _w is the weight of the heaviest chain aviable from point_i.
        # iterate over point_j again, adding to iset all the jset points only from _w weighted point_js.
        for j in range(i + 1, len(points)):
            point_j = (points[j][0], points[j][1])
            # avoid sub-chains by passing points that already in iset.
            if point_j in iset:
                continue
            if point_j[0] > point_i[0] and point_j[1] > point_i[1]:
                if (value := point_set_map.get(points[j])) is not None:
                    (jweight, jset) = value
                    if _w > jweight:
                        continue

                    # point_j meets the conditions...
                    # insert to iset all the jset point.
                    iset.update(jset)
                    iset.add((point_i))

        if not iset:
            # this point has no comparable bigger points.
            # update the map to hold an iset of itself with its weight.
            iset = set()
            iset.add(point_i)
            point_set_map[point_i] = (weights_map.get(point_i, 1), iset)
        else:
            # update the map: update with new iset and iweight = _w + weight of point_i
            point_set_map[point_i] = (_w + weights_map.get(point_i, 1), iset)

        # for printing percentage...
        if perc % 10 == 0:
            if perc != 100:
                print(f"{perc}%", end=" ", flush=True)
            else:
                print(f"{perc}%")

    # reutrn the map
    return point_set_map

def update_points_weights(sorted_points, point_set_map, weights_map, start_w, W):
    """given the points and the point to sets map, this function increase weights of points that starts max weight chain (weighted less than W) by 1.
    call this function after get_set_of_points_on_heaviest_chains_chains() to increase the overall weights.
    can be called in iterations to maximize weights."""

    print("2. update_points_weights()")

    ending_weight = start_w

    # A flag to identify a increased happend. True will cause another iteration.
    increased = False
    visited = set()
    for point in sorted_points:
        # if point is on a maxed weight chain, continue.
        if point in visited:
            continue

        _w, point_set = point_set_map[point]

        # add point_set to the visited set.
        visited.update(point_set)

        # check if point weight can be increased.
        can_increased_by = W - _w
        if can_increased_by > 0:
            # update the weight of the point.
            weights_map[point] = weights_map.get(point, 1) + 1
            # update the weight of the heavies chain starting from point, for next iterations.
            point_set_map[point] = (_w + 1, point_set)

            # for printing.
            ending_weight += 1

            # identify a increased happend. this will cause another iteration.
            increased = True

    return ending_weight, increased

# Sum of squared weights functions
def get_set_of_points_on_heaviest_chains_squared(points, weights_map):
    """returns a map: KEY = point (x,y) , VALUE = ( _w , set) : _w is the weight of heaviest chain starting at point. 
        set is a set of all points on _w weight chains starting from point."""

    print("1. get_set_of_points_on_heaviest_chains_squared()")

    # KEY = point , VALUE = (_w, set()) . this is the returned map.
    point_set_map = {}

    """
    Iterate from end to start.
    for each point (point_i) calculate the following:
    1. iset = set of all points that are on one of the heaviest weight chains starting from point_i.
    2. _w = weight of the heaviest chains.
    
    """
    for i in range(len(points) - 1, -1, -1):
        # for printing.
        perc = 100 - (i / len(points)) * 100

        point_i = (points[i][0], points[i][1])
        iset = set()
        _w = 0

        # find the max _w belongs to any comparable bigger point (point_j)
        for j in range(i + 1, len(points)):
            point_j = (points[j][0], points[j][1])
            if point_j[0] > point_i[0] and point_j[1] > point_i[1]:
                # point_j is a comparable bigger point!
                if (value := point_set_map.get(points[j])) is not None:
                    # point_j has been found.
                    # value = (point_j_w ,jset)
                    jweight = value[0]
                    _w = max(_w, jweight)

        # now _w is the weight of the heaviest chain aviable from point_i.
        # iterate over point_j again, adding to iset all the jset points only from _w weighted point_js.
        for j in range(i + 1, len(points)):
            point_j = (points[j][0], points[j][1])
            # avoid sub-chains by passing points that already in iset.
            if point_j in iset:
                continue
            if point_j[0] > point_i[0] and point_j[1] > point_i[1]:
                if (value := point_set_map.get(points[j])) is not None:
                    (jweight, jset) = value
                    curr_weight = weights_map.get(point_i, 1)

                    # skip all point_j that are not _w weighted or will not get impacted if point_i weight will be increased by one.
                    if jweight + (curr_weight + 1) ** 2 < _w:
                        continue

                    # insert to iset all the jset point.
                    iset.update(jset)
                    iset.add((point_i))

        if not iset:
            # this point has no comparable bigger points.
            # update the map to hold an iset of itself with its weight.
            iset = set()
            iset.add(point_i)
            point_set_map[point_i] = (weights_map.get(point_i, 1) ** 2, iset)
        else:
            # update the map: update with new iset and iweight = _w + weight of point_i
            point_set_map[point_i] = (_w + weights_map.get(point_i, 1) ** 2, iset)

        # for printing percentage...
        if perc % 10 == 0:
            if perc != 100:
                print(f"{perc}%", end=" ", flush=True)
            else:
                print(f"{perc}%")

    # reutrn the map
    return point_set_map

def update_points_weights_squared(sorted_points, point_set_map, weights_map, start_w, W):
    """given the points and the point to sets map, this function increase weights of points that starts max weight chain (weighted less than W) by 1.
        call this function after get_set_of_points_on_heaviest_chains_chains_squared() to increase the overall weights.
        can be called in iterations to maximize weights."""


    print("2. update_points_weights_squared()")

    ending_weight = start_w

    # A flag to identify a increased happend. True will cause another iteration.
    increased = False
    visited = set()
    for point in sorted_points:
        # if point is on a maxed weight chain, continue.
        if point in visited:
            continue

        _w, point_set = point_set_map[point]

        curr_weight = weights_map.get(point, 1)
        new_w = _w -  curr_weight ** 2  + (curr_weight + 1) ** 2

        # add point_set to the visited set.
        visited.update(point_set)

        # check if point weight can be increased.
        can_increased_by = W - new_w
        if can_increased_by > 0:
            # update the weight of the point.

            # need to increase point weight by 1.
            # increase _w by one each iteration.
            weights_map[point] = weights_map.get(point, 1) + 1
            # update the weight of the heavies chain starting from point, for next iterations.
            point_set_map[point] = (new_w, point_set)

            # for printing.
            ending_weight += 1

            # identify a increased happend. this will cause another iteration.
            increased = True

    return ending_weight, increased

# Program flow functions
def maximum_weight_iteration(points, weights_map, start_w, W, squared):
    if squared:
        point_set_map = get_set_of_points_on_heaviest_chains_squared(points, weights_map)
        end_w, is_increased = update_points_weights_squared(
            points, point_set_map, weights_map, start_w, W
        )
        return end_w, is_increased
    else:
        point_set_map = get_set_of_points_on_heaviest_chains(points, weights_map)
        end_w, is_increased = update_points_weights(
            points, point_set_map, weights_map, start_w, W
        )
        return end_w, is_increased

def run(points, squared=False):
    # points = normalize_points(points)
    if not squared: 
        print("---------Running with sum of regular weights...") 
    else: 
        print("---------Running with sum of squared weights..")
    # ______  saved results for printing.
    all_weights = []

    # ------- initalizition.
    run_iteration = True
    weights_map = {}
    # point_set_map = {}
    start_w = len(points)
    i = 0

    # start time
    start_time = time.time()

    # ------- run iterations.
    print(f"Calculating W = ", end=" ", flush=True)
    first_W, _ = get_W(points, weights_map,squared)
    print(f"{first_W}")
    while run_iteration:
        i += 1
        print(f"running iteration [{i}]")
        end_w, run_iteration = maximum_weight_iteration(
            points, weights_map, start_w, first_W, squared=squared
        )

        # for printing.
        all_weights.append((start_w, end_w))
        start_w = end_w

        # check if excceding max run time.
        time_passed = time.time() - start_time
        if time_passed / 60 > MAX_RUNTIME_MINUTES:
            run_iteration = False
        # else: run_iteration = False
        print(f"time_passed: {time_passed:.0f} sec")

    # _____________________________________________________________________

    last_W, weighted_points = get_W(points, weights_map,squared)
    total_time = time.time() - start_time
    print("Iterations Completed!")
    print(f"running time : {total_time}")


    # ___________ print results
    print(
        f"In {i} iterations: total weight increase: {all_weights[0][0]} -> {all_weights[-1][1]}",flush=True
    )
    
    print(
        f"Max chain weight (W) sanity check : first iteration = {first_W}, after {i} iterations = {last_W} "
    ,flush=True)

    return weights_map

# Main
if __name__ == "__main__":

    MAX_RUNTIME_MINUTES = 1  # in minutes.

    strart_time = time.time()

    sheets = get_sheets_names()
    for i,sheet_name in enumerate(sheets):
        points,original_points = get_points_from_sheet(sheet_name)
        print(f"Reading data from Sheet: {sheet_name}")
        weighted_points_map = run(points)
        squared_weighted_points_map = run(points, squared=True)
        # write_output(original_points,weighted_points_map, squared_weighted_points_map,sheet_name)
    total_time_minutes = (time.time() - strart_time) / 60

    print(f"Total time: {total_time_minutes:.3f} minutes.")
